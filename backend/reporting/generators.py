"""
ParamX Hunter - Report Generators
PDF (ReportLab), HTML (Jinja2), Excel (openpyxl)
"""

import os
from datetime import datetime
from pathlib import Path

from jinja2 import Environment, BaseLoader
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from backend.database.models import Endpoint, Parameter, Scan


# ── Data Loader ────────────────────────────────────────────────────────────────

async def _load_scan_data(db: AsyncSession, scan_id: str) -> dict:
    scan_res = await db.execute(select(Scan).where(Scan.id == scan_id))
    scan = scan_res.scalar_one_or_none()

    params_res = await db.execute(
        select(Parameter).where(Parameter.scan_id == scan_id).order_by(Parameter.risk_level)
    )
    params = params_res.scalars().all()

    eps_res = await db.execute(
        select(Endpoint).where(Endpoint.scan_id == scan_id)
    )
    endpoints = eps_res.scalars().all()

    by_risk: dict[str, int] = {}
    by_type: dict[str, int] = {}
    for p in params:
        r = str(p.risk_level)
        t = str(p.param_type)
        by_risk[r] = by_risk.get(r, 0) + 1
        by_type[t] = by_type.get(t, 0) + 1

    return {
        "scan": scan,
        "parameters": params,
        "endpoints": endpoints,
        "generated_at": datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC"),
        "summary": {
            "total_endpoints": len(endpoints),
            "total_parameters": len(params),
            "sensitive": sum(1 for p in params if p.is_sensitive),
            "hidden": sum(1 for p in params if p.is_hidden),
            "critical": by_risk.get("critical", 0),
            "high": by_risk.get("high", 0),
            "apis": sum(1 for e in endpoints if e.is_api),
            "websockets": sum(1 for e in endpoints if e.is_websocket),
            "by_risk": by_risk,
            "by_type": by_type,
        },
    }


# ── HTML Report ────────────────────────────────────────────────────────────────

HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>ParamX Hunter Report — {{ scan.name }}</title>
<style>
  :root { --cyan: #06b6d4; --bg: #050e1a; --card: #0d1520; --border: #1e2a3a; }
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: 'Segoe UI', system-ui, sans-serif; background: var(--bg); color: #e2e8f0; padding: 40px; }
  h1 { font-size: 2rem; color: var(--cyan); margin-bottom: 4px; }
  h2 { font-size: 1.1rem; color: #94a3b8; margin: 32px 0 12px; border-bottom: 1px solid var(--border); padding-bottom: 8px; }
  .meta { color: #64748b; font-size: 0.85rem; margin-bottom: 32px; font-family: monospace; }
  .grid { display: grid; grid-template-columns: repeat(4, 1fr); gap: 16px; margin-bottom: 32px; }
  .card { background: var(--card); border: 1px solid var(--border); border-radius: 10px; padding: 20px; }
  .card .num { font-size: 2rem; font-weight: 700; color: var(--cyan); font-family: monospace; }
  .card .label { font-size: 0.7rem; color: #64748b; text-transform: uppercase; letter-spacing: 0.1em; margin-top: 4px; }
  table { width: 100%; border-collapse: collapse; font-size: 0.82rem; }
  th { background: #07101c; padding: 10px 12px; text-align: left; color: #64748b; font-size: 0.7rem; text-transform: uppercase; letter-spacing: 0.1em; border-bottom: 1px solid var(--border); }
  td { padding: 9px 12px; border-bottom: 1px solid var(--border); font-family: monospace; color: #94a3b8; }
  tr:hover td { background: #0d1520; }
  .badge { display: inline-block; padding: 2px 8px; border-radius: 4px; font-size: 0.7rem; font-weight: 600; text-transform: uppercase; }
  .critical { background: rgba(239,68,68,.15); color: #f87171; }
  .high     { background: rgba(249,115,22,.15); color: #fb923c; }
  .medium   { background: rgba(251,191,36,.15); color: #fbbf24; }
  .low      { background: rgba(52,211,153,.15);  color: #34d399; }
  .info     { background: rgba(96,165,250,.15);  color: #60a5fa; }
  .sensitive { color: #fbbf24; }
  footer { margin-top: 40px; padding-top: 20px; border-top: 1px solid var(--border); color: #334155; font-size: 0.75rem; text-align: center; }
</style>
</head>
<body>
<h1>⚡ ParamX Hunter</h1>
<p class="meta">
  Scan: {{ scan.name }} &nbsp;|&nbsp;
  Generated: {{ generated_at }} &nbsp;|&nbsp;
  Status: {{ scan.status }}
</p>

<div class="grid">
  <div class="card"><div class="num">{{ summary.total_endpoints }}</div><div class="label">Endpoints</div></div>
  <div class="card"><div class="num">{{ summary.total_parameters }}</div><div class="label">Parameters</div></div>
  <div class="card"><div class="num">{{ summary.sensitive }}</div><div class="label">Sensitive</div></div>
  <div class="card"><div class="num" style="color:#ef4444">{{ summary.critical }}</div><div class="label">Critical Risk</div></div>
  <div class="card"><div class="num">{{ summary.hidden }}</div><div class="label">Hidden Fields</div></div>
  <div class="card"><div class="num">{{ summary.apis }}</div><div class="label">APIs Discovered</div></div>
  <div class="card"><div class="num">{{ summary.websockets }}</div><div class="label">WebSockets</div></div>
  <div class="card"><div class="num" style="color:#f97316">{{ summary.high }}</div><div class="label">High Risk</div></div>
</div>

<h2>Parameters</h2>
<table>
  <thead>
    <tr>
      <th>Name</th><th>Type</th><th>Source</th><th>Risk</th><th>Flags</th><th>Freq.</th>
    </tr>
  </thead>
  <tbody>
    {% for p in parameters %}
    <tr>
      <td>{{ p.name }}</td>
      <td>{{ p.param_type }}</td>
      <td>{{ p.source }}</td>
      <td><span class="badge {{ p.risk_level }}">{{ p.risk_level }}</span></td>
      <td>
        {% if p.is_sensitive %}<span class="sensitive">⚠ sensitive</span>{% endif %}
        {% if p.is_hidden %}🕵 hidden{% endif %}
      </td>
      <td>{{ p.frequency }}</td>
    </tr>
    {% endfor %}
  </tbody>
</table>

<h2>Endpoints</h2>
<table>
  <thead>
    <tr><th>Path</th><th>Method</th><th>Status</th><th>Type</th><th>Framework</th></tr>
  </thead>
  <tbody>
    {% for ep in endpoints %}
    <tr>
      <td>{{ ep.path }}</td>
      <td>{{ ep.method }}</td>
      <td>{{ ep.status_code or "—" }}</td>
      <td>
        {% if ep.is_graphql %}GraphQL
        {% elif ep.is_websocket %}WebSocket
        {% elif ep.is_api %}REST API
        {% else %}Page{% endif %}
      </td>
      <td>{{ ep.framework_detected or "—" }}</td>
    </tr>
    {% endfor %}
  </tbody>
</table>

<footer>
  Generated by ParamX Hunter — For authorized security testing only.
</footer>
</body>
</html>
"""


async def generate_html_report(
    db: AsyncSession,
    scan_id: str,
    output_path: str,
    title: str,
    include_values: bool = False,
) -> None:
    data = await _load_scan_data(db, scan_id)
    env = Environment(loader=BaseLoader())
    tmpl = env.from_string(HTML_TEMPLATE)
    html = tmpl.render(**data)
    Path(output_path).write_text(html, encoding="utf-8")


# ── PDF Report ─────────────────────────────────────────────────────────────────

async def generate_pdf_report(
    db: AsyncSession,
    scan_id: str,
    output_path: str,
    title: str,
    include_values: bool = False,
) -> None:
    """Generate PDF via ReportLab."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    )

    data = await _load_scan_data(db, scan_id)
    scan = data["scan"]
    params = data["parameters"]
    endpoints = data["endpoints"]
    summary = data["summary"]

    doc = SimpleDocTemplate(output_path, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    mono = ParagraphStyle("mono", fontName="Courier", fontSize=8, leading=10)
    title_style = ParagraphStyle("title", fontSize=22, textColor=colors.HexColor("#06b6d4"),
                                 fontName="Helvetica-Bold", spaceAfter=4)
    h2_style = ParagraphStyle("h2", fontSize=13, textColor=colors.HexColor("#94a3b8"),
                               fontName="Helvetica-Bold", spaceBefore=18, spaceAfter=8)

    RISK_COLORS = {
        "critical": colors.HexColor("#ef4444"),
        "high": colors.HexColor("#f97316"),
        "medium": colors.HexColor("#fbbf24"),
        "low": colors.HexColor("#34d399"),
        "info": colors.HexColor("#60a5fa"),
    }

    story = []
    story.append(Paragraph("⚡ ParamX Hunter", title_style))
    story.append(Paragraph(f"Scan: <b>{scan.name}</b> — {data['generated_at']}", styles["Normal"]))
    story.append(Spacer(1, 0.5*cm))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#1e2a3a")))
    story.append(Spacer(1, 0.3*cm))

    # Summary table
    sum_data = [
        ["Endpoints", "Parameters", "Sensitive", "Critical", "High", "APIs"],
        [
            str(summary["total_endpoints"]),
            str(summary["total_parameters"]),
            str(summary["sensitive"]),
            str(summary["critical"]),
            str(summary["high"]),
            str(summary["apis"]),
        ]
    ]
    sum_table = Table(sum_data, colWidths=[2.8*cm]*6)
    sum_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#07101c")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#64748b")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("FONTNAME", (0, 1), (-1, 1), "Courier-Bold"),
        ("TEXTCOLOR", (0, 1), (-1, 1), colors.HexColor("#06b6d4")),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#1e2a3a")),
        ("ROWBACKGROUNDS", (0, 1), (-1, 1), [colors.HexColor("#0d1520")]),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(sum_table)
    story.append(Spacer(1, 0.5*cm))

    # Parameters table
    story.append(Paragraph("Parameters", h2_style))
    p_rows = [["Name", "Type", "Source", "Risk", "Flags"]]
    for p in params[:500]:  # cap at 500 for PDF size
        flags = []
        if p.is_sensitive:
            flags.append("sensitive")
        if p.is_hidden:
            flags.append("hidden")
        p_rows.append([
            Paragraph(p.name[:50], mono),
            Paragraph(str(p.param_type).replace("_", " "), mono),
            Paragraph(p.source, mono),
            Paragraph(str(p.risk_level).upper(), mono),
            Paragraph(", ".join(flags), mono),
        ])

    p_table = Table(p_rows, colWidths=[4.5*cm, 3*cm, 3*cm, 2*cm, 2.5*cm])
    p_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#07101c")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#64748b")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 7),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#1e2a3a")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#050e1a"), colors.HexColor("#0d1520")]),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(p_table)

    doc.build(story)


# ── Excel Report ───────────────────────────────────────────────────────────────

async def generate_excel_report(
    db: AsyncSession,
    scan_id: str,
    output_path: str,
    include_values: bool = False,
) -> None:
    """Generate Excel report with multiple sheets."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    data = await _load_scan_data(db, scan_id)
    wb = openpyxl.Workbook()

    HDR_FILL = PatternFill("solid", fgColor="07101C")
    HDR_FONT = Font(color="64748B", bold=True, name="Consolas", size=9)
    CYAN_FONT = Font(color="06B6D4", name="Consolas", size=9)
    MONO_FONT = Font(name="Consolas", size=9)
    RISK_FILLS = {
        "critical": PatternFill("solid", fgColor="1A0505"),
        "high": PatternFill("solid", fgColor="1A0A05"),
        "medium": PatternFill("solid", fgColor="1A1505"),
        "low": PatternFill("solid", fgColor="051A0A"),
        "info": PatternFill("solid", fgColor="050A1A"),
    }

    def add_sheet(name: str, headers: list[str], rows: list[list]) -> None:
        ws = wb.create_sheet(name)
        for col, hdr in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=hdr)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal="left")
        for r, row in enumerate(rows, 2):
            for col, val in enumerate(row, 1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.font = MONO_FONT
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].auto_size = True

    # Summary sheet
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.sheet_view.showGridLines = False
    summary_rows = [
        ["Scan Name", data["scan"].name],
        ["Generated At", data["generated_at"]],
        ["Total Endpoints", data["summary"]["total_endpoints"]],
        ["Total Parameters", data["summary"]["total_parameters"]],
        ["Sensitive Parameters", data["summary"]["sensitive"]],
        ["Hidden Fields", data["summary"]["hidden"]],
        ["Critical Risk", data["summary"]["critical"]],
        ["High Risk", data["summary"]["high"]],
        ["APIs Discovered", data["summary"]["apis"]],
        ["WebSockets", data["summary"]["websockets"]],
    ]
    for row_data in summary_rows:
        ws_sum.append(row_data)

    # Parameters sheet
    param_headers = ["Name", "Type", "Source", "Method", "Risk Level",
                     "Sensitive", "Hidden", "Frequency", "Risk Tags", "First Seen"]
    if include_values:
        param_headers.insert(1, "Value")

    param_rows = []
    for p in data["parameters"]:
        row = [p.name, str(p.param_type), p.source, str(p.method or ""),
               str(p.risk_level), str(p.is_sensitive), str(p.is_hidden),
               p.frequency, ", ".join(p.risk_tags or []),
               p.first_seen.strftime("%Y-%m-%d %H:%M")]
        if include_values:
            row.insert(1, (p.value or "")[:200])
        param_rows.append(row)
    add_sheet("Parameters", param_headers, param_rows)

    # Endpoints sheet
    ep_headers = ["URL", "Method", "Status", "Is API", "Is GraphQL",
                  "Is WebSocket", "Framework", "First Seen"]
    ep_rows = [
        [ep.url, str(ep.method), str(ep.status_code or ""), str(ep.is_api),
         str(ep.is_graphql), str(ep.is_websocket), ep.framework_detected or "",
         ep.first_seen.strftime("%Y-%m-%d %H:%M")]
        for ep in data["endpoints"]
    ]
    add_sheet("Endpoints", ep_headers, ep_rows)

    wb.save(output_path)
